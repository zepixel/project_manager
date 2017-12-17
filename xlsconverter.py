# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from jinja2 import Template, Environment, PackageLoader, select_autoescape
import os


class xlsx_doc:
	
	def __init__(self,name):
		self.name = name
		self.first_row = 0
		self.first_col = 0
		
		# Structures de données
		#self.document : workbook. cf: self.load().

		self.element = {} 			# dico d'un élement. Cree a partir du header du doc.
		self.liste_elements = [] 	# Ensemble des elements d'une feuille.
		self.feuille = {"NOM":"","ELEMENTS": self.liste_elements}  
		self.classeur = []

		self.mapping = {}  			# coordonnées de mapping d'une liste de projet de session sur les colonnes du fichier xlsx.
	



	def load(self):

		#--- Listing et selection des fichiers xlsx présents dans le dossier
		fichiers=[fichier for fichier in os.listdir(".") if (".xlsx" in fichier or ".xls" in fichier) ]
		
		for index,fichier in enumerate(fichiers):
			print ("[%s] %s" %(index,fichier))
		
		self.document = load_workbook(fichiers[int(input("\nEntrez le Numero de votre fichier.\n>> "))])
		
		print("Fichier chargé.\n")


		#--- Première Ligne et première colonne du tableau ?
		self.first_row = int(input("Entrez le numero de la première ligne utile du tableau:\n>> "))
		self.first_col = int(input("Entrez le numero de la première colonne utile du tableau:\n>> "))

		#--- Y a-t-il un en-tête dans le tableau ?
		self.header = True    # on va dire que oui.


		'''self.header = input("\nLe tableau possède t-il un entête ?\n1. Oui\n2. Non\n>> ")

		if (self.header == "1"):
			self.header = bool(1)
		else:
			self.header = bool(0)
	
		print(self.header)
		'''



		# Attribution des clés du dico d'un element en fonction de l'entete du tableau s'il y en à un.
		if self.header:
			for sheet in self.document:
			
				for ligne in sheet.iter_rows(min_row = self.first_row, min_col=self.first_col, max_row = self.first_row):
				
					for key in ligne:
						self.element.update({key.value:""})
				break
			
		else:
			print("pas d'entete")

	
	
		# Attribution des valeurs aux clés du dictionnaire d'éléments en fonction des lignes du tableau
		for sheet in self.document:
			
			self.feuille["NOM"] = sheet.title
			
			for ligne in sheet.iter_rows(min_row = self.first_row + 1 , min_col=self.first_col):
				index_case_ligne = 0    
			
				for key in self.element.keys():
					self.element[key]=ligne[index_case_ligne].value
					index_case_ligne = index_case_ligne + 1

					#with open("./Web/lol.txt","a") as fichier:     //Test error UTF 8
					#	fichier.write(str(self.element[key]))
				
				self.liste_elements.append(self.element.copy())
				
				self.feuille["ELEMENTS"]= self.liste_elements.copy()

			self.classeur.append(self.feuille.copy())
			self.liste_elements = []
			

		print ("Structure de donnée créee.")   
		print("\n")




	def map(self, table):

		# ToDo: Mise en place d'un système de template de mapping en fonction des formats des xlsx des écoles
		# ex: If template_mapping =="esra" self.mapping == {esra_xlsx:esra_projectList  etc}

		automap = True
		table = table
		print ("Les colonnes dans le fichier xlsx sont les suivantes :\n")
		out = ""

		for key_element in self.element.keys():
			out = (out + "| " + key_element +" " )
		
		input (out + "|")

		for key_element in self.element.keys():
			
			if automap == True:
			
				if key_element == "NOM":
					self.mapping.update({"NOM":"ETUDIANTS"})
					continue

				if key_element == "NOTE":	
					self.mapping.update({"NOTE":"NOTE"})
					continue
				
				if key_element == "SECTION":	
					self.mapping.update({"SECTION":"none"})
					continue
				
				if key_element == "CLASSE":	
					self.mapping.update({"CLASSE":"none"})
					continue

				if key_element == "PRENOM":	
					self.mapping.update({"PRENOM":"none"})
					continue

				if key_element == "OBSERVATION":
					self.mapping.update({"OBSERVATION":"COMMENTAIRE"})

				else:

					print("\n- Quelle valeur voulez vous attribuer à la colonne : " + key_element + " ?\n")
					choix_map =[]
					for index, key_table in enumerate(table[0].keys()):
						print(str(index) + ". " + key_table)
						choix_map.append(key_table)

					self.mapping.update({key_element:choix_map[int(input(">> "))]})
					print (choix_map)
					

			else:

				print("\n- Quelle valeur voulez vous attribuer à la colonne : " + key_element + " ?\n")
				choix_map =[]
				for index, key_table in enumerate(table[0].keys()):
					print(str(index) + ". " + key_table)
					choix_map.append(key_table)

				self.mapping.update({key_element:choix_map[int(input(">> "))]})
				print (choix_map)

			
			input (self.mapping)


		




class web_renderer:

	def __init__(self,name,document):
		self.name = name
		#self.env = env
		self.document = document
		self.title = input("\nQuel sera le titre de la page Web ?\n>> ") or " "
		self.description = input("\nQuelle description souhaitez-vous afficher sur la page ?\n>> ") or " "

		
		#--- Styles de couleurs de la page
		self.colorstyle_default = {
		"nom" : "Style par defaut",
		"page_bg" : "#F8E6E0",
		"header_bg" : "#00538c",
		"sections_bg" : "#0062a6",
		"sections_hover" : "#00538c",
		"notice_bg" : "#F2F2F2",
		"tab_header" : "#4CAF50",
		"tab_header_hover" : "#3d8d40" 
		}

		self.colorstyle_1 = {
		"nom" : "Nocturne",
		"page_bg" : "#F8E6E0",
		"header_bg" : "#192B38",
		"sections_bg" : "#22424B",
		"sections_hover" : "#2E5458",
		"notice_bg" : "#98C7A8",
		"tab_header" : "#4E7367",
		"tab_header_hover" : "#6B9E8E" 
		}

		self.colorstyle_2 = {
		"nom": "Palette neutre",
		"page_bg" : "#ECE3D9",
		"header_bg" : "#5C6E6B",
		"sections_bg" : "#8F8781",
		"sections_hover" : "#B2A9A1",
		"notice_bg" : "#D4CFC5",
		"tab_header" : "#B0AAA4",
		"tab_header_hover" : "#BDB6AF" 
		}

		self.colorstyle_3 = {
		"nom": "Black Cherry Mocha",
		"page_bg" : "#FFF6D9",
		"header_bg" : "#705B35",
		"sections_bg" : "#C7B07B",
		"sections_hover" : "#D6BD85",
		"notice_bg" : "#E8D9AC",
		"tab_header" : "#9E814B",
		"tab_header_hover" : "#C29E5C" 
		}

		# Stockage dans la liste de style
		self.style_list = [self.colorstyle_default, self.colorstyle_1, self.colorstyle_2, self.colorstyle_3]



	def style_select(self):

		# --- Selection du style ---#
		print("\nChoisissez un style de couleurs: \n")
		index = 0
		for index,style in enumerate(self.style_list):
			print( str(index) + ". " + style["nom"] + "\n")
		self.choix_style = int(input(">> ") or "1")



	def load_templates(self):
		# Chargement Templates
		# INDEX.HTML
		self.template_html_file = open("./templates/template_index.html", "r" )
		self.template_css_file = open('./templates/template_style.css',"r")
		
		self.indexhtml = Template(self.template_html_file.read())
		# STYLE.CSS
		self.stylecss = Template(self.template_css_file.read())
		print("templates chargés")

		self.template_html_file.close()
		self.template_css_file.close()


	def render(self):

		#Création des fichiers de sortie
		self.file_index = "./Web/Index.html"
		self.file_stylecss = "./Web/style.css"

		if not os.path.exists("./Web"):
			os.makedirs("./Web")

		with open(self.file_index,"w+") as fichier:
			output = self.indexhtml.render(classeur= self.document.classeur, titre_page=self.title , description_page =self.description, en_tete=self.document.header, feuille=self.document.feuille)
			fichier.write(output)

		with open(self.file_stylecss,"w+") as fichier:
			fichier.write(self.stylecss.render(choix_style = self.style_list[self.choix_style - 1] ))

		input("\nFichier converti. Mettez en ligne le contenu du dossier nommé 'Web'.")





#--- Chargement de l'environement Jinja   ---> !!! BUG !!! cree une loop indésirable dans l'execution du programme.
'''env = Environment(
loader=PackageLoader('JINJA_XLS_CONVERTER', 'templates'),
autoescape=select_autoescape(['html', 'xml'])
)'''